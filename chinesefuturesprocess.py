import os
import pandas as pd
import xlsxwriter
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill, GradientFill
from datetime import datetime
from xlrd import open_workbook
from pathlib import Path
import win32com.client
import xlwings as xw
from PIL import ImageGrab

x = datetime.now()
print(x.strftime("%d%b%y"))
month = x.strftime("%b")
year = x.strftime("%y")

#setting the cell border style
thin = Side(border_style="thin", color="000000")
thick = Side(border_style="thick", color="000000")
double = Side(border_style="double", color="000000")

def createheader():
    #setting up the header for table
    sheet['B1'] = str('http://quote.eastmoney.com/qihuo/ZCM.html')
    sheet['B1'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, shrink_to_fit=False, indent=0)
    sheet.merge_cells('B1:E1')
    sheet['B2'] = str('Chinese Futures ZCM')
    sheet['B2'].font = Font(bold=True)
    sheet['B2'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, shrink_to_fit=True, indent=0)
    sheet['B2'].border = Border(top=thick, left=thick, right=thick, bottom=thick)
    sheet.merge_cells('B2:E2')
    sheet['B3'] = str('Date')
    sheet['B3'].font = Font(bold=True)
    sheet['B3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, shrink_to_fit=True, indent=0)
    sheet['B3'].fill = PatternFill("solid", fgColor = "00008000")
    sheet['B3'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    sheet['C3'] = str('Time/Session')
    sheet['C3'].font = Font(bold=True)
    sheet['C3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, shrink_to_fit=True, indent=0)
    sheet['C3'].fill = PatternFill("solid", fgColor = "00008000")
    sheet['C3'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    sheet['D3'] = str('Futures Price')
    sheet['D3'].font = Font(bold=True)
    sheet['D3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, shrink_to_fit=True, indent=0)
    sheet['D3'].fill = PatternFill("solid", fgColor = "00008000")
    sheet['D3'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    sheet['E3'] = str('Changes')
    sheet['E3'].font = Font(bold=True)
    sheet['E3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, shrink_to_fit=True, indent=0)
    sheet['E3'].fill = PatternFill("solid", fgColor = "00008000")
    sheet['E3'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    wb.save(path)

def deletesummary():
    wbdest = openpyxl.load_workbook(path)
    wsdest = wbdest.active
    maxrow = wsdest.max_row
    wsdest.delete_rows(maxrow,1)
    print("deleting row: ", maxrow)
    wbdest.save(path)

#define excel name and its path
excelfilename = 'Chinese Futures Prices ZCM.xlsx'
path = r'C:\Users\asus\Desktop\KCH\Automation\ChineseFutures\%s' %excelfilename

#checking if excel file already exists, create one if it doesn't exist
if os.path.isfile(path):
    print("Excel file has been made, let's process the data!")
else:
    print("Creating the excel file...")
    writer = pd.ExcelWriter(path)
    writer.save()
    #change worksheet name
    wb = openpyxl.load_workbook(path, read_only=False)
    sheet = wb["Sheet1"]
    sheet.title = "Chinese Futures %s %s" %(month, year)
    sheet.column_dimensions['B'].width = 14
    sheet.column_dimensions['C'].width = 14
    sheet.row_dimensions[1].height = 32
    createheader()

#checking if worksheet already exists, create one if it doesn't exist
wb = openpyxl.load_workbook(path, read_only=False)
sheettitle = "Chinese Futures %s %s" %(month, year)

if sheettitle in wb.sheetnames:
    print("Sheet: ", sheettitle, "already exist")
    wb = openpyxl.load_workbook(path, read_only=False)
    wb.active = 0
    for sheet in wb:
        if sheet.title == sheettitle:
            sheet.sheet_view.tabSelected = True
        else:
            sheet.sheet_view.tabSelected = False
        createheader()
else:
    print("We are creating new worksheet for you named:", sheettitle)
    sheet = wb.create_sheet(sheettitle,0)
    sheet = wb[sheettitle]
    sheet.column_dimensions['B'].width = 14
    sheet.column_dimensions['C'].width = 14
    #sheet.row_dimension[1].height = 32
    wb.active = 0
    for sheet in wb:
        if sheet.title == sheettitle:
            sheet.sheet_view.tabSelected = True
        else:
            sheet.sheet_view.tabSelected = False
        createheader()
    
    



def datainput():
    #opening source excel file
    Filename = r'C:\Users\asus\Desktop\KCH\Automation\ChineseFutures\ChineseFutures' + x.strftime("%d%m%y %I%p") + '.xlsx'
    wbsource = openpyxl.load_workbook(Filename)
    wssource = wbsource.worksheets[0]

    maxrowsource = wssource.max_row
    maxcolumnsource = wssource.max_column

    print("max row in source file:", maxrowsource)
    print("max column in source file:", maxcolumnsource)
    datacount = maxrowsource - 1

    #opening destination excel file
    wbdest = openpyxl.load_workbook(path)
    wsdest = wbdest.active

    maxrowdest = wsdest.max_row
    maxcolumndest = wsdest.max_column

    print("max row in destination file:", maxrowdest)
    print("max column in destination file:", maxcolumndest)
    
    wslast = wb.worksheets[1]
    maxrowprev = wslast.max_row
    
    if maxrowdest == 3:
        wsdest.cell(row=(maxrowdest+1), column=2).value = x.strftime("%d/%b/%y")
        wsdest.cell(row=(maxrowdest+1), column=2).font = Font(bold = True)
        wsdest.cell(row=(maxrowdest+1), column=2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, shrink_to_fit=True, indent=0)
        wsdest.cell(row=(maxrowdest+1), column=2).border = Border(top=thin, left=thin, right=thin, bottom=thin)

        i = 0
        for i in range(0, datacount):
            wsdest.cell(row=(maxrowdest+1), column=3).value = wssource.cell(row=i+2, column=2).value
            wsdest.cell(row=(maxrowdest+1), column=3).font = Font(bold = True)
            wsdest.cell(row=(maxrowdest+1), column=3).border = Border(top=thin, left=thin, right=thin, bottom=thin)
            wsdest.cell(row=(maxrowdest+1), column=3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, shrink_to_fit=True, indent=0)
            wsdest.cell(row=(maxrowdest+1), column=4).value = wssource.cell(row=i+2, column=3).value
            wsdest.cell(row=(maxrowdest+1), column=4).font = Font(bold = True)
            wsdest.cell(row=(maxrowdest+1), column=4).border = Border(top=thin, left=thin, right=thin, bottom=thin)
            wsdest.cell(row=(maxrowdest+1), column=4).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, shrink_to_fit=True, indent=0)
            wsdest.cell(row=(maxrowdest+1), column=5).value = (wsdest.cell(row=(maxrowdest+1), column =4).value) - (wslast.cell(row=(maxrowprev-1), column=4).value)
            wsdest.cell(row=(maxrowdest+1), column=5).font = Font(bold = True)
            wsdest.cell(row=(maxrowdest+1), column=5).border = Border(top=thin, left=thin, right=thin, bottom=thin)
            wsdest.cell(row=(maxrowdest+1), column=5).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, shrink_to_fit=True, indent=0)
            wsdest.cell(row=(maxrowdest+1), column=5).fill = PatternFill("solid", fgColor = "00C0C0C0")
            maxrowdest = maxrowdest + 1
 
        wbdest.save(path)
    
    else:
        deletesummary()
        wsdest.unmerge_cells(start_row=maxrowdest, start_column=2, end_row=maxrowdest, end_column=3)
        wsdest.cell(row=(maxrowdest), column=2).value = x.strftime("%d/%b/%y")
        wsdest.cell(row=(maxrowdest), column=2).font = Font(bold = True)
        wsdest.cell(row=(maxrowdest), column=2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, shrink_to_fit=True, indent=0)
        wsdest.cell(row=(maxrowdest), column=2).border = Border(top=thin, left=thin, right=thin, bottom=thin)

        i = 0
        for i in range(0, datacount):
            wsdest.cell(row=(maxrowdest), column=3).value = wssource.cell(row=i+2, column=2).value
            wsdest.cell(row=(maxrowdest), column=3).font = Font(bold = True)
            wsdest.cell(row=(maxrowdest), column=3).border = Border(top=thin, left=thin, right=thin, bottom=thin)
            wsdest.cell(row=(maxrowdest), column=3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, shrink_to_fit=True, indent=0)
            wsdest.cell(row=(maxrowdest), column=4).value = wssource.cell(row=i+2, column=3).value
            wsdest.cell(row=(maxrowdest), column=4).font = Font(bold = True)
            wsdest.cell(row=(maxrowdest), column=4).border = Border(top=thin, left=thin, right=thin, bottom=thin)
            wsdest.cell(row=(maxrowdest), column=4).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, shrink_to_fit=True, indent=0)
            wsdest.cell(row=(maxrowdest), column=5).value = float(wsdest.cell(row=maxrowdest, column=4).value) - float(wsdest.cell(row=maxrowdest-1, column=4).value)
            wsdest.cell(row=(maxrowdest), column=5).border = Border(top=thin, left=thin, right=thin, bottom=thin)
            wsdest.cell(row=(maxrowdest), column=5).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, shrink_to_fit=True, indent=0)
            wsdest.cell(row=(maxrowdest), column=5).fill = PatternFill("solid", fgColor = "00C0C0C0")
            if wsdest.cell(row=(maxrowdest), column=5).value < 0:
                wsdest.cell(row=(maxrowdest), column=5).font = Font(bold = True, color = "00FF0000")
            elif wsdest.cell(row=(maxrowdest), column=5).value == 0:
                wsdest.cell(row=(maxrowdest), column=5).font = Font(bold = True)
            else:
                wsdest.cell(row=(maxrowdest), column=5).font = Font(bold = True, color = "00008000")
            maxrowdest = maxrowdest + 1
 
        wbdest.save(path)
        
datainput()

def summarize():
    wbdest = openpyxl.load_workbook(path)
    wsdest = wbdest.active

    data=pd.read_excel(path)

    maxrowdest = wsdest.max_row
    maxcolumndest = wsdest.max_column
    print('sampe sini maxrowdest:', maxrowdest)
    wsdest.cell(row=(maxrowdest+1), column=2).value = "Chinese Futures Average :"
    wsdest.cell(row=(maxrowdest+1), column=2).font = Font(bold = True)
    wsdest.cell(row=(maxrowdest+1), column=2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, shrink_to_fit=True, indent=0)
    wsdest.cell(row=(maxrowdest+1), column=2).border = Border(top=thin, left=thin, right=thin, bottom=thin)
    mergerange = 'B%s:C%s' %(maxrowdest+1, maxrowdest+1)
    wsdest.merge_cells(mergerange)
    
    sum = 0
    for i in range(4, maxrowdest+1):
        sum = sum + float(wsdest.cell(row=i, column=4).value)
        i = i + 1
    countdata = maxrowdest - 3
    average = sum/countdata
    
    wsdest.cell(row=(maxrowdest+1), column=4).value = average
    wsdest.cell(row=(maxrowdest+1), column=4).font = Font(bold = True)
    wsdest.cell(row=(maxrowdest+1), column=4).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, shrink_to_fit=True, indent=0)
    wsdest.cell(row=(maxrowdest+1), column=4).border = Border(top=thin, left=thin, right=thin, bottom=thin)
    
    wbdest.save(path)
    
summarize()

#produceimage()